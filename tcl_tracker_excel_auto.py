#!/usr/bin/env python3
"""
TCL Electronics KSA — Automated Excel Training Tracker
Generates a fully self-contained .xlsx workbook with:
  - Promoter entry sheet with auto-calculated points & levels
  - Monthly training plan with dropdown status & progress bar
  - Competition scoring sheet with weighted rubric formulas
  - Auto-sorted leaderboard via formulas
  - Charts that update as data is entered
"""

from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, FormulaRule, CellIsRule
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / f"TCL_Automated_Tracker_{date.today().strftime('%Y_%m_%d')}.xlsx"

# ── Colours ───────────────────────────────────────────────────────────────────
RED        = "C8102E"
DARK       = "1A1A2E"
WHITE      = "FFFFFF"
LIGHT_RED  = "FAD7DC"
LIGHT_GREY = "F5F5F5"
MID_GREY   = "D9D9D9"
GREEN      = "1A7A3C"
LIGHT_GRN  = "DDFFD6"
GOLD       = "FFD700"
SILVER     = "C0C0C0"
BRONZE     = "CD7F32"
YELLOW     = "FFF3CC"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=DARK, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def border(color=MID_GREY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, cols, bg=RED, fg=WHITE, height=24):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill(bg); cell.font = fnt(bold=True, color=fg)
        cell.alignment = ctr(); cell.border = border()
    ws.row_dimensions[row].height = height

def banner(ws, title, sub, cols="A:H"):
    ws.merge_cells(f"A1:{cols.split(':')[1]}1")
    c = ws["A1"]
    c.value = title; c.fill = fill(RED)
    c.font = fnt(bold=True, color=WHITE, size=16); c.alignment = ctr()
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f"A2:{cols.split(':')[1]}2")
    c = ws["A2"]
    c.value = sub; c.fill = fill(DARK)
    c.font = fnt(color=WHITE, size=10); c.alignment = ctr()
    ws.row_dimensions[2].height = 20

def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — PROMOTER TRACKER
# ─────────────────────────────────────────────────────────────────────────────
def sheet_promoters(wb):
    ws = wb.create_sheet("Promoter Tracker")
    ws.sheet_view.showGridLines = False
    banner(ws, "TCL PROMOTER TRACKER — KSA",
           f"Auto-calculates Points, Level & Status  |  Updated: {date.today().strftime('%d %B %Y')}", "A:J")

    # Instructions row
    ws.merge_cells("A3:J3")
    ws["A3"].value = (
        "📝  HOW TO USE:  Enter promoter data in the yellow cells. "
        "Points, Level, and Status are calculated automatically."
    )
    ws["A3"].fill = fill(YELLOW)
    ws["A3"].font = fnt(size=10, italic=True, color=DARK)
    ws["A3"].alignment = lft()
    ws.row_dimensions[3].height = 22

    # Header
    headers = ["ID", "Name", "Store / Region",
               "Bronze\nQuiz", "Silver\nQuiz", "Gold\nQuiz",
               "Competition\nPoints", "Total\nPoints", "Level", "Badge"]
    hdr(ws, 4, headers, height=36)

    # Points per quiz level (hidden reference — column L)
    ws["L1"] = "Bronze pts"; ws["M1"] = 10
    ws["L2"] = "Silver pts"; ws["M2"] = 20
    ws["L3"] = "Gold pts";   ws["M3"] = 35

    # 20 promoter rows
    for i in range(1, 21):
        row = i + 4
        bg = LIGHT_GREY if i % 2 == 0 else WHITE

        # ID auto-fill
        id_cell = ws.cell(row=row, column=1, value=f"P{i:03d}")
        id_cell.fill = fill(bg); id_cell.font = fnt(bold=True, color=RED)
        id_cell.alignment = ctr(); id_cell.border = border()

        # Name & Store — yellow input cells
        for col in (2, 3):
            c = ws.cell(row=row, column=col, value="")
            c.fill = fill(YELLOW); c.alignment = lft(); c.border = border()

        # Quiz columns (D–F) — dropdown: Pending / Pass / Fail
        for col in (4, 5, 6):
            c = ws.cell(row=row, column=col, value="Pending")
            c.fill = fill(YELLOW); c.alignment = ctr(); c.border = border()

        # Competition points — manual yellow input
        cp = ws.cell(row=row, column=7, value=0)
        cp.fill = fill(YELLOW); cp.alignment = ctr(); cp.border = border()

        # Total points formula
        # Bronze pass=10, Silver pass=20, Gold pass=35 + comp points
        total_formula = (
            f'=IF(D{row}="Pass",$M$1,0)'
            f'+IF(E{row}="Pass",$M$2,0)'
            f'+IF(F{row}="Pass",$M$3,0)'
            f'+G{row}'
        )
        tp = ws.cell(row=row, column=8, value=total_formula)
        tp.fill = fill(bg); tp.font = fnt(bold=True, color=DARK)
        tp.alignment = ctr(); tp.border = border()

        # Level formula
        level_formula = (
            f'=IF(H{row}>=150,"Champion",'
            f'IF(H{row}>=100,"Gold",'
            f'IF(H{row}>=50,"Silver",'
            f'IF(H{row}>0,"Bronze","—"))))'
        )
        lv = ws.cell(row=row, column=9, value=level_formula)
        lv.fill = fill(bg); lv.alignment = ctr(); lv.border = border()

        # Badge formula
        badge_formula = (
            f'=IF(I{row}="Champion","🏆",'
            f'IF(I{row}="Gold","🥇",'
            f'IF(I{row}="Silver","🥈",'
            f'IF(I{row}="Bronze","🥉","—"))))'
        )
        bv = ws.cell(row=row, column=10, value=badge_formula)
        bv.fill = fill(bg); bv.font = fnt(size=14)
        bv.alignment = ctr(); bv.border = border()

        ws.row_dimensions[row].height = 22

    # ── Dropdown validation for quiz columns ──
    dv_quiz = DataValidation(
        type="list", formula1='"Pass,Fail,Pending"',
        allow_blank=False, showDropDown=False
    )
    dv_quiz.sqref = f"D5:F24"
    ws.add_data_validation(dv_quiz)

    # ── Conditional formatting ──
    # Level column — colour by level
    for text, color in [("Champion", RED), ("Gold", GOLD), ("Silver", SILVER), ("Bronze", BRONZE)]:
        ws.conditional_formatting.add(
            "I5:I24",
            CellIsRule(operator="equal", formula=[f'"{text}"'],
                       fill=PatternFill("solid", fgColor=color),
                       font=Font(bold=True, color=DARK, name="Calibri"))
        )
    # Quiz pass = green, fail = red
    ws.conditional_formatting.add(
        "D5:F24",
        CellIsRule(operator="equal", formula=['"Pass"'],
                   fill=PatternFill("solid", fgColor=LIGHT_GRN))
    )
    ws.conditional_formatting.add(
        "D5:F24",
        CellIsRule(operator="equal", formula=['"Fail"'],
                   fill=PatternFill("solid", fgColor=LIGHT_RED))
    )
    # Data bar on total points
    ws.conditional_formatting.add(
        "H5:H24",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num",   end_value=150,
                    color=RED)
    )

    # Summary box (right side)
    ws.merge_cells("L5:N5")
    ws["L5"].value = "TEAM SUMMARY"; ws["L5"].fill = fill(DARK)
    ws["L5"].font = fnt(bold=True, color=WHITE); ws["L5"].alignment = ctr()

    summaries = [
        ("Total Promoters",  '=COUNTA(B5:B24)'),
        ("Avg Points",       '=IFERROR(AVERAGE(H5:H24),0)'),
        ("Champions",        '=COUNTIF(I5:I24,"Champion")'),
        ("Gold",             '=COUNTIF(I5:I24,"Gold")'),
        ("Silver",           '=COUNTIF(I5:I24,"Silver")'),
        ("Bronze",           '=COUNTIF(I5:I24,"Bronze")'),
        ("Quizzes Passed",   '=COUNTIF(D5:F24,"Pass")'),
    ]
    for r, (label, formula) in enumerate(summaries, 6):
        ws.cell(row=r, column=12, value=label).fill = fill(LIGHT_GREY)
        ws.cell(row=r, column=12).font = fnt(size=10)
        ws.cell(row=r, column=12).border = border()
        ws.cell(row=r, column=12).alignment = lft()
        val = ws.cell(row=r, column=13, value=formula)
        val.fill = fill(YELLOW); val.font = fnt(bold=True, color=RED, size=12)
        val.alignment = ctr(); val.border = border()
        ws.row_dimensions[r].height = 22

    col_w(ws, [7, 22, 22, 10, 10, 10, 12, 10, 12, 8, 4, 18, 10])


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — TRAINING PLAN
# ─────────────────────────────────────────────────────────────────────────────
TASKS = [
    # (week, label, id, day, task)
    (1,"Foundation (Jun 1–7)","w1d1","Mon Jun 1","Full-body strength – 3×10 squats, push-ups, rows"),
    (1,"Foundation (Jun 1–7)","w1d2","Tue Jun 2","30-min steady-state cardio (run/cycle)"),
    (1,"Foundation (Jun 1–7)","w1d3","Wed Jun 3","Core & mobility – planks, hip flexor stretches"),
    (1,"Foundation (Jun 1–7)","w1d4","Thu Jun 4","Upper body – bench press, overhead press, pull-ups"),
    (1,"Foundation (Jun 1–7)","w1d5","Fri Jun 5","Lower body – deadlifts, lunges, calf raises"),
    (1,"Foundation (Jun 1–7)","w1d6","Sat Jun 6","Active recovery – 45-min walk or yoga"),
    (1,"Foundation (Jun 1–7)","w1d7","Sun Jun 7","Rest day"),
    (2,"Building (Jun 8–14)","w2d1","Mon Jun 8","Full-body strength – increase weight 5%"),
    (2,"Building (Jun 8–14)","w2d2","Tue Jun 9","Interval cardio – 6×3-min hard / 2-min easy"),
    (2,"Building (Jun 8–14)","w2d3","Wed Jun 10","Core circuit – 4 rounds: hollow hold, dead bug, RKC"),
    (2,"Building (Jun 8–14)","w2d4","Thu Jun 11","Upper body – add 1 set per exercise"),
    (2,"Building (Jun 8–14)","w2d5","Fri Jun 12","Lower body – add 1 set per exercise"),
    (2,"Building (Jun 8–14)","w2d6","Sat Jun 13","Hike or long bike ride (60 min)"),
    (2,"Building (Jun 8–14)","w2d7","Sun Jun 14","Rest day"),
    (3,"Intensity (Jun 15–21)","w3d1","Mon Jun 15","Full-body HIIT – 20 min Tabata"),
    (3,"Intensity (Jun 15–21)","w3d2","Tue Jun 16","Tempo run – 5 km at comfortably hard pace"),
    (3,"Intensity (Jun 15–21)","w3d3","Wed Jun 17","Yoga / deep stretch (45 min)"),
    (3,"Intensity (Jun 15–21)","w3d4","Thu Jun 18","Push day – chest, shoulders, triceps 4×8"),
    (3,"Intensity (Jun 15–21)","w3d5","Fri Jun 19","Pull day – back, biceps, rear delts 4×8"),
    (3,"Intensity (Jun 15–21)","w3d6","Sat Jun 20","Leg day – heavy squats & RDLs 4×6"),
    (3,"Intensity (Jun 15–21)","w3d7","Sun Jun 21","Rest day"),
    (4,"Peak (Jun 22–28)","w4d1","Mon Jun 22","Max-effort full-body circuit – 5 rounds"),
    (4,"Peak (Jun 22–28)","w4d2","Tue Jun 23","Long run – 8 km at easy pace"),
    (4,"Peak (Jun 22–28)","w4d3","Wed Jun 24","Core & stability – single-leg work, pallof press"),
    (4,"Peak (Jun 22–28)","w4d4","Thu Jun 25","Upper body strength test – max reps benchmark"),
    (4,"Peak (Jun 22–28)","w4d5","Fri Jun 26","Lower body strength test – 1RM squat & deadlift"),
    (4,"Peak (Jun 22–28)","w4d6","Sat Jun 27","Active recovery swim or cycle (45 min)"),
    (4,"Peak (Jun 22–28)","w4d7","Sun Jun 28","Rest day"),
    (5,"Recovery (Jun 29–30)","w5d1","Mon Jun 29","Light full-body movement – 50% intensity"),
    (5,"Recovery (Jun 29–30)","w5d2","Tue Jun 30","End-of-month review + mobility session"),
]

def sheet_training_plan(wb):
    ws = wb.create_sheet("Training Plan")
    ws.sheet_view.showGridLines = False
    banner(ws, "JUNE 2026 — MONTHLY TRAINING PLAN",
           "Select status from dropdown. Progress updates automatically.", "A:G")

    # Progress summary row
    ws.merge_cells("A3:D3")
    ws["A3"].value = "MONTHLY PROGRESS"
    ws["A3"].fill = fill(DARK); ws["A3"].font = fnt(bold=True, color=WHITE)
    ws["A3"].alignment = ctr(); ws.row_dimensions[3].height = 24

    ws.merge_cells("E3:G3")
    progress_formula = '=COUNTIF(E6:E35,"✓ Done")&" / 30 tasks ("&TEXT(COUNTIF(E6:E35,"✓ Done")/30,"0%")&" complete)"'
    ws["E3"].value = progress_formula
    ws["E3"].fill = fill(YELLOW); ws["E3"].font = fnt(bold=True, color=RED, size=12)
    ws["E3"].alignment = ctr(); ws.row_dimensions[3].height = 24

    # Data bar across row 4
    ws.merge_cells("A4:G4")
    ws["A4"].value = '=COUNTIF(E6:E35,"✓ Done")/30'
    ws["A4"].number_format = "0%"
    ws["A4"].fill = fill(LIGHT_GRN); ws["A4"].font = fnt(bold=True, color=GREEN, size=12)
    ws["A4"].alignment = ctr(); ws.row_dimensions[4].height = 28

    ws.conditional_formatting.add(
        "A4:A4",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num",   end_value=1,
                    color=GREEN)
    )

    hdr(ws, 5, ["Week", "Phase", "ID", "Day", "Status", "Task", "Notes"])

    prev_week = None
    for i, (week, label, tid, day, task) in enumerate(TASKS):
        row = i + 6
        bg = LIGHT_GREY if week % 2 == 0 else WHITE

        # Week number (merge across same week)
        wc = ws.cell(row=row, column=1,
                     value=week if week != prev_week else "")
        wc.fill = fill(LIGHT_RED if week % 2 != 0 else MID_GREY)
        wc.font = fnt(bold=True, color=RED if week != prev_week else LIGHT_GREY, size=14)
        wc.alignment = ctr(); wc.border = border()

        # Phase label
        pc = ws.cell(row=row, column=2, value=label if week != prev_week else "")
        pc.fill = fill(bg); pc.font = fnt(size=9, italic=True)
        pc.alignment = lft(); pc.border = border()

        # ID
        idc = ws.cell(row=row, column=3, value=tid)
        idc.fill = fill(bg); idc.font = fnt(bold=True, color=DARK, size=9)
        idc.alignment = ctr(); idc.border = border()

        # Day
        dc = ws.cell(row=row, column=4, value=day)
        dc.fill = fill(bg); dc.alignment = ctr(); dc.border = border()

        # Status — dropdown
        sc = ws.cell(row=row, column=5, value="○ To Do")
        sc.fill = fill(YELLOW); sc.alignment = ctr(); sc.border = border()
        sc.font = fnt(bold=True)

        # Task
        tc = ws.cell(row=row, column=6, value=task)
        tc.fill = fill(bg); tc.alignment = lft(); tc.border = border()

        # Notes — free text
        nc = ws.cell(row=row, column=7, value="")
        nc.fill = fill(WHITE); nc.alignment = lft(); nc.border = border()

        ws.row_dimensions[row].height = 22
        prev_week = week

    # Status dropdown
    dv = DataValidation(
        type="list",
        formula1='"✓ Done,○ To Do,— Skip"',
        allow_blank=False, showDropDown=False
    )
    dv.sqref = "E6:E35"
    ws.add_data_validation(dv)

    # Conditional formatting on status
    ws.conditional_formatting.add(
        "A6:G35",
        FormulaRule(formula=['$E6="✓ Done"'],
                    fill=PatternFill("solid", fgColor=LIGHT_GRN))
    )
    ws.conditional_formatting.add(
        "A6:G35",
        FormulaRule(formula=['$E6="— Skip"'],
                    fill=PatternFill("solid", fgColor=MID_GREY))
    )

    col_w(ws, [7, 22, 8, 14, 12, 48, 20])


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — COMPETITION SCORER
# ─────────────────────────────────────────────────────────────────────────────
def sheet_competition(wb):
    ws = wb.create_sheet("Competition Scorer")
    ws.sheet_view.showGridLines = False
    banner(ws, "COMPETITION SCORING SHEET",
           "Rubric: Knowledge 35%  ·  Creativity 25%  ·  Selling 40%  |  Score 1–4 per criterion", "A:J")

    # Rubric reference
    ws["L1"] = "Knowledge weight"; ws["M1"] = 0.35
    ws["L2"] = "Creativity weight"; ws["M2"] = 0.25
    ws["L3"] = "Selling weight";    ws["M3"] = 0.40

    ws.merge_cells("A3:J3")
    ws["A3"].value = (
        "📝  Enter competition name in B4. Score each promoter 1–4 in columns D, E, F. "
        "Weighted score, rank and award are calculated automatically."
    )
    ws["A3"].fill = fill(YELLOW); ws["A3"].font = fnt(size=10, italic=True)
    ws["A3"].alignment = lft(); ws.row_dimensions[3].height = 22

    # Competition name input
    ws.merge_cells("A4:C4")
    ws["A4"].value = "Competition Name:"
    ws["A4"].fill = fill(DARK); ws["A4"].font = fnt(bold=True, color=WHITE)
    ws["A4"].alignment = lft()
    ws.merge_cells("D4:J4")
    ws["D4"].value = "C7L Mastery Excellence — Phase 1"
    ws["D4"].fill = fill(YELLOW); ws["D4"].font = fnt(bold=True, color=RED, size=12)
    ws["D4"].alignment = lft()
    ws.row_dimensions[4].height = 28

    hdr(ws, 5, [
        "Rank", "Promoter Name", "Store",
        "Knowledge\n/4 (35%)", "Creativity\n/4 (25%)", "Selling\n/4 (40%)",
        "Weighted\nScore", "Score\n%", "Result", "Award\n(SAR)"
    ], height=40)

    for i in range(1, 21):
        row = i + 5
        bg = LIGHT_GREY if i % 2 == 0 else WHITE

        # Name & store — yellow input
        for col in (2, 3):
            c = ws.cell(row=row, column=col, value="")
            c.fill = fill(YELLOW); c.alignment = lft(); c.border = border()

        # Score inputs 1–4
        for col in (4, 5, 6):
            c = ws.cell(row=row, column=col, value="")
            c.fill = fill(YELLOW); c.alignment = ctr(); c.border = border()

        # Weighted score formula
        ws_formula = (
            f'=IFERROR(D{row}*$M$1+E{row}*$M$2+F{row}*$M$3,"")'
        )
        wsc = ws.cell(row=row, column=7, value=ws_formula)
        wsc.fill = fill(bg); wsc.font = fnt(bold=True)
        wsc.number_format = "0.000"; wsc.alignment = ctr(); wsc.border = border()

        # Pct
        pct = ws.cell(row=row, column=8,
                      value=f'=IFERROR(G{row}/4,"")' )
        pct.fill = fill(bg); pct.number_format = "0%"
        pct.alignment = ctr(); pct.border = border()

        # Result (top 3 label)
        rank_f = (
            f'=IFERROR(IF(RANK(G{row},G6:G25,0)=1,"🥇 1st",'
            f'IF(RANK(G{row},G6:G25,0)=2,"🥈 2nd",'
            f'IF(RANK(G{row},G6:G25,0)=3,"🥉 3rd",""))),"—")'
        )
        rc = ws.cell(row=row, column=9, value=rank_f)
        rc.fill = fill(bg); rc.alignment = ctr(); rc.border = border()

        # Award
        award_f = (
            f'=IFERROR(IF(RANK(G{row},G6:G25,0)=1,1000,'
            f'IF(RANK(G{row},G6:G25,0)=2,800,'
            f'IF(RANK(G{row},G6:G25,0)=3,600,""))),"—")'
        )
        ac = ws.cell(row=row, column=10, value=award_f)
        ac.fill = fill(bg); ac.number_format = '#,##0 "SAR"'
        ac.alignment = ctr(); ac.border = border()

        # Rank (hidden col A for sorting reference)
        rank_num = ws.cell(row=row, column=1,
                           value=f'=IFERROR(RANK(G{row},G6:G25,0),"")')
        rank_num.fill = fill(bg); rank_num.alignment = ctr()
        rank_num.border = border(); rank_num.font = fnt(bold=True, color=RED)

        ws.row_dimensions[row].height = 22

    # Score input validation 1–4
    dv_score = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="4",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Score",
        error="Please enter a number between 1 and 4."
    )
    dv_score.sqref = "D6:F25"
    ws.add_data_validation(dv_score)

    # Colour top 3 rows
    for rank_val, bg_color in [(1, LIGHT_RED), (2, YELLOW), (3, LIGHT_GRN)]:
        ws.conditional_formatting.add(
            f"A6:J25",
            FormulaRule(
                formula=[f'$A6={rank_val}'],
                fill=PatternFill("solid", fgColor=bg_color),
                font=Font(bold=True, name="Calibri")
            )
        )

    # Data bar on weighted score
    ws.conditional_formatting.add(
        "G6:G25",
        DataBarRule(start_type="num", start_value=1,
                    end_type="num",   end_value=4, color=RED)
    )

    col_w(ws, [7, 24, 22, 13, 13, 12, 10, 9, 10, 12, 4, 18, 10])


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 — LEADERBOARD (auto-updated via formulas)
# ─────────────────────────────────────────────────────────────────────────────
def sheet_leaderboard(wb):
    ws = wb.create_sheet("Leaderboard")
    ws.sheet_view.showGridLines = False
    banner(ws, f"PROMOTER LEADERBOARD — {date.today().strftime('%B %Y').upper()}",
           "Auto-sorted by Total Points from the Promoter Tracker sheet", "A:G")

    ws.merge_cells("A3:G3")
    ws["A3"].value = "ℹ️  This leaderboard reads live from the Promoter Tracker sheet. It updates as you enter data."
    ws["A3"].fill = fill(YELLOW); ws["A3"].font = fnt(size=10, italic=True)
    ws["A3"].alignment = lft(); ws.row_dimensions[3].height = 22

    hdr(ws, 4, ["Rank", "Promoter", "Store", "Points", "Level", "Badge", "Progress Bar"])

    medals = {1: "🥇", 2: "🥈", 3: "🥉"}
    for i in range(1, 21):
        row = i + 4
        bg = LIGHT_RED if i <= 3 else (LIGHT_GREY if i % 2 == 0 else WHITE)

        # Medal
        medal = ws.cell(row=row, column=1, value=medals.get(i, str(i)))
        medal.fill = fill(bg); medal.font = fnt(bold=True, size=14 if i<=3 else 11)
        medal.alignment = ctr(); medal.border = border()

        # LARGE formula to get Nth largest promoter name by points
        name_f = (
            f'=IFERROR(INDEX(\'Promoter Tracker\'!B$5:B$24,'
            f'MATCH(LARGE(\'Promoter Tracker\'!H$5:H$24,{i}),'
            f'\'Promoter Tracker\'!H$5:H$24,0)),"")'
        )
        nc = ws.cell(row=row, column=2, value=name_f)
        nc.fill = fill(bg); nc.alignment = lft(); nc.border = border()
        if i <= 3: nc.font = fnt(bold=True)

        store_f = (
            f'=IFERROR(INDEX(\'Promoter Tracker\'!C$5:C$24,'
            f'MATCH(LARGE(\'Promoter Tracker\'!H$5:H$24,{i}),'
            f'\'Promoter Tracker\'!H$5:H$24,0)),"")'
        )
        sc2 = ws.cell(row=row, column=3, value=store_f)
        sc2.fill = fill(bg); sc2.alignment = lft(); sc2.border = border()

        pts_f = f'=IFERROR(LARGE(\'Promoter Tracker\'!H$5:H$24,{i}),"")'
        pc = ws.cell(row=row, column=4, value=pts_f)
        pc.fill = fill(bg); pc.font = fnt(bold=True, color=RED if i<=3 else DARK, size=13 if i<=3 else 11)
        pc.alignment = ctr(); pc.border = border()

        level_f = (
            f'=IFERROR(INDEX(\'Promoter Tracker\'!I$5:I$24,'
            f'MATCH(LARGE(\'Promoter Tracker\'!H$5:H$24,{i}),'
            f'\'Promoter Tracker\'!H$5:H$24,0)),"")'
        )
        lc = ws.cell(row=row, column=5, value=level_f)
        lc.fill = fill(bg); lc.alignment = ctr(); lc.border = border()

        badge_f = (
            f'=IFERROR(INDEX(\'Promoter Tracker\'!J$5:J$24,'
            f'MATCH(LARGE(\'Promoter Tracker\'!H$5:H$24,{i}),'
            f'\'Promoter Tracker\'!H$5:H$24,0)),"")'
        )
        bc = ws.cell(row=row, column=6, value=badge_f)
        bc.fill = fill(bg); bc.font = fnt(size=16)
        bc.alignment = ctr(); bc.border = border()

        # Progress bar cell (data bar via cond. formatting)
        prog_f = f'=IFERROR(LARGE(\'Promoter Tracker\'!H$5:H$24,{i})/150,"")'
        pbc = ws.cell(row=row, column=7, value=prog_f)
        pbc.fill = fill(bg); pbc.number_format = "0%"
        pbc.alignment = ctr(); pbc.border = border()

        ws.row_dimensions[row].height = 26 if i <= 3 else 22

    ws.conditional_formatting.add(
        "G5:G24",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num",   end_value=1, color=RED)
    )

    col_w(ws, [8, 26, 24, 10, 14, 8, 18])


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 — INSTRUCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def sheet_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)   # first sheet
    ws.sheet_view.showGridLines = False
    banner(ws, "TCL ELECTRONICS KSA — AUTOMATED TRAINING TRACKER",
           "Complete guide to using this workbook", "A:E")

    steps = [
        ("📋  PROMOTER TRACKER", [
            "1. Go to the 'Promoter Tracker' sheet",
            "2. Enter Name and Store in the yellow cells (rows 5–24)",
            "3. Change quiz status using the dropdown: Pass / Fail / Pending",
            "4. Enter competition points earned in column G",
            "5. Points, Level, and Badge are calculated automatically",
        ]),
        ("🗓  TRAINING PLAN", [
            "1. Go to the 'Training Plan' sheet",
            "2. Select a status from the dropdown in column E: ✓ Done / ○ To Do / — Skip",
            "3. The progress bar and counter update automatically",
            "4. Add notes in column G for any session",
        ]),
        ("🏆  COMPETITION SCORER", [
            "1. Go to the 'Competition Scorer' sheet",
            "2. Enter the competition name in cell D4",
            "3. Enter promoter names and stores in columns B and C",
            "4. Score each criterion 1–4 (Knowledge, Creativity, Selling)",
            "5. Weighted score, ranking, and award (SAR) appear automatically",
            "6. Top 3 are highlighted in red / yellow / green",
        ]),
        ("📊  LEADERBOARD", [
            "1. Go to the 'Leaderboard' sheet",
            "2. It reads live from Promoter Tracker — no manual update needed",
            "3. Refreshes automatically when you change data in Promoter Tracker",
        ]),
        ("💡  TIPS", [
            "• Save frequently — Ctrl+S",
            "• Use Ctrl+Z to undo any mistakes",
            "• Yellow cells = input  |  White/grey cells = calculated automatically",
            "• Do NOT edit the calculated columns (Points, Level, Badge, Score)",
            "• To add more than 20 promoters, copy and extend the row formulas",
        ]),
    ]

    row = 4
    for title, points in steps:
        ws.merge_cells(f"A{row}:E{row}")
        tc = ws.cell(row=row, column=1, value=title)
        tc.fill = fill(DARK); tc.font = fnt(bold=True, color=WHITE, size=13)
        tc.alignment = lft(); ws.row_dimensions[row].height = 28
        row += 1
        for pt in points:
            ws.merge_cells(f"A{row}:E{row}")
            pc = ws.cell(row=row, column=1, value=pt)
            pc.fill = fill(LIGHT_GREY if row % 2 == 0 else WHITE)
            pc.font = fnt(size=11); pc.alignment = lft()
            ws.row_dimensions[row].height = 22
            row += 1
        row += 1

    col_w(ws, [6, 30, 30, 20, 20])


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def generate():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_instructions(wb)
    sheet_promoters(wb)
    sheet_training_plan(wb)
    sheet_competition(wb)
    sheet_leaderboard(wb)

    # Set active sheet to instructions on open
    wb.active = wb["Instructions"]

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    generate()
