"""
Fill KSA_April TCL Trainer Template with real April 2026 data.
Reads apr_records.pkl (82 sessions) and writes to a new Excel file.
"""

import pickle
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import copy

TEMPLATE = "/root/.claude/uploads/82218592-4f36-5112-b730-0955daaf23f9/777e64de-KSA_April__TCL_Trainer_Template_v3_FINAL_6_1.xlsx"
OUTPUT = "/home/user/greet/KSA_April_TCL_Trainer_Filled_2026.xlsx"
APR_PKL = "/home/user/greet/apr_records.pkl"
ALL_PKL = "/home/user/greet/tv_training_parsed.pkl"

# ── Content → Topic mapping ──────────────────────────────────────────────────
TOPIC_MAP_INSTORE = {
    "main features":          "L3 Instore: C7L Key Selling Points",
    "what is sqd":            "L3 Instore: SQD Technology Explanation",
    "a-to-a":                 "L2 Instore: Up-Selling Technique",
    "sell talk":              "L2 Instore: Up-Selling Technique",
    "product knowledge":      "L2 Instore: TCL Full Line-Up Overview",
    "full features":          "L2 Instore: TCL Full Line-Up Overview",
    "c7l features":           "L3 Instore: C7L Key Selling Points",
    "c8l features":           "L3 Instore: C8L Key Selling Points",
    "x11l features":          "L3 Instore: X11L Key Selling Points",
    "sqd technology":         "L3 Instore: SQD Technology Explanation",
    "up selling":             "L2 Instore: Up-Selling Technique",
    "benchmark":              "L2 Instore: TCL Benchmark",
    "picture quality":        "L1 Instore: Picture Quality",
    "sound quality":          "L1 Instore: Sound Quality",
    "gaming":                 "L1 Instore: Gaming Features",
    "ports":                  "L1 Instore: TV Ports & Inputs",
    "connectivity":           "L1 Instore: Smart System & Connectivity",
}

TOPIC_MAP_CLASS = {
    "2026 intro":             "Level 2: TCL Line-Up & Culture",
    "main features":          "Level 3: SQD MiniLED",
    "product knowledge":      "Level 2: TCL Line-Up & Culture",
    "full features":          "Level 2: TCL Line-Up & Culture",
    "npi":                    "NPI: New Product Introduction",
    "new product":            "NPI: New Product Introduction",
    "advanced":               "Level 4: Advanced (4K)",
    "4k":                     "Level 4: Advanced (4K)",
    "soundbar":               "Level 4: Advanced (Soundbar)",
    "basics":                 "Level 1: TV Basics",
    "sqd":                    "Level 3: SQD MiniLED",
}

TOPIC_MAP_ONLINE = {
    "product knowledge":      "Level 2: TCL Line-Up & Culture",
    "main features":          "Level 3: SQD MiniLED",
    "2026 intro":             "Level 2: TCL Line-Up & Culture",
    "sqd":                    "Level 3: SQD MiniLED",
}

def map_topic(content, method):
    key = content.lower().strip()
    if method == "Instore":
        for k, v in TOPIC_MAP_INSTORE.items():
            if k in key:
                return v
        return "L2 Instore: TCL Full Line-Up Overview"   # default instore
    elif method == "Class":
        for k, v in TOPIC_MAP_CLASS.items():
            if k in key:
                return v
        return "Level 2: TCL Line-Up & Culture"          # default class
    else:  # Online
        for k, v in TOPIC_MAP_ONLINE.items():
            if k in key:
                return v
        return "Level 2: TCL Line-Up & Culture"


def map_method(raw):
    m = raw.strip().lower()
    if m in ("on-site", "on-Site", "instore"):
        return "Instore"
    if m == "class":
        return "Class"
    if m == "online":
        return "Online"
    return "Instore"


def parse_score(score_str):
    """Convert '89%' or '0.89' to float 0.89, or None."""
    if not score_str or score_str in ("N/A", "n/a", ""):
        return None
    s = str(score_str).strip().rstrip("%")
    try:
        v = float(s)
        if v > 1:
            v /= 100
        return round(v, 4)
    except ValueError:
        return None


def parse_nps(nps_str):
    """Convert '78%' → 7.8  or '9.5' → 9.5,  or None."""
    if not nps_str or nps_str in ("N/A", "n/a", ""):
        return None
    s = str(nps_str).strip()
    if "%" in s:
        try:
            return round(float(s.rstrip("%")) / 10, 1)
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date_display(date_str):
    """Keep date as string like '1-Apr' → '01 Apr 2026'."""
    if not date_str:
        return ""
    parts = date_str.split("-")
    if len(parts) == 2:
        day, month = parts
        return f"{int(day):02d} {month} 2026"
    return date_str


def clear_rows(ws, start_row, end_row=None):
    """Clear cell values from data rows, skipping merged cells."""
    from openpyxl.cell.cell import MergedCell
    max_row = end_row or ws.max_row
    for row in ws.iter_rows(min_row=start_row, max_row=max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def write_row(ws, row_num, values, style_fill=None):
    """Write a list of values to a row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = val
        if style_fill:
            cell.fill = style_fill


# ── Load data ────────────────────────────────────────────────────────────────
with open(APR_PKL, "rb") as f:
    apr = pickle.load(f)

# Load all rows for May plan
with open(ALL_PKL, "rb") as f:
    all_data = pickle.load(f)

all_rows = all_data["rows"]
# Jun-dated rows represent "next month" (May plan) sessions
jun_rows = [r for r in all_rows if len(r) > 5 and "Jun" in str(r[5])]

# ── Compute summary stats ────────────────────────────────────────────────────
total_sessions   = len(apr)
class_sessions   = sum(1 for r in apr if r.get("method","") == "Class")
online_sessions  = sum(1 for r in apr if r.get("method","") == "Online")
instore_sessions = sum(1 for r in apr if r.get("method","") == "On-Site")
total_participants = sum(r.get("participants_num", 0) or 0 for r in apr)

scored = [r for r in apr if r.get("score_num") is not None]
pass_count = sum(1 for r in scored if (r.get("score_num") or 0) >= 90)
exam_pass_rate = round(pass_count / len(scored), 2) if scored else 0.0

print(f"Sessions: {total_sessions}  (Class={class_sessions}, Online={online_sessions}, Instore={instore_sessions})")
print(f"Participants: {total_participants}")
print(f"Scored sessions: {len(scored)},  Pass (>=90%): {pass_count},  Rate: {exam_pass_rate:.0%}")

# ── Open template ────────────────────────────────────────────────────────────
print("\nLoading template…")
wb = load_workbook(TEMPLATE)

# ════════════════════════════════════════════════════════════════════════════
# 1.  SUMMARY SHEET — update Achieved column only
# ════════════════════════════════════════════════════════════════════════════
ws_sum = wb["📋 Summary"]

# Achieved cells are in column C, rows 13-19
ws_sum["C13"] = total_sessions        # Total Training Sessions
ws_sum["C14"] = class_sessions        # Class
ws_sum["C15"] = online_sessions       # Online
ws_sum["C16"] = instore_sessions      # Instore
ws_sum["C17"] = total_participants    # Total Trainees Reached
ws_sum["C18"] = exam_pass_rate        # Exam Pass Rate (decimal)
ws_sum["C19"] = 2                     # Content pieces (keep existing)

print("✓ Summary sheet updated")

# ════════════════════════════════════════════════════════════════════════════
# 2.  SESSION LOG — clear sample rows and write all 82 sessions
# ════════════════════════════════════════════════════════════════════════════
ws_log = wb["📅 Session Log"]

# Clear existing data rows (5 and beyond)
clear_rows(ws_log, start_row=5, end_row=ws_log.max_row)

yellow_fill = PatternFill("solid", fgColor="FFFF99")

for idx, r in enumerate(apr, 1):
    row_num = idx + 4   # data starts at row 5

    method  = map_method(r.get("method", "On-Site"))
    content = r.get("content", "")
    topic   = map_topic(content, method)
    model   = r.get("model", "")
    score_raw = r.get("score_num")
    score_d = round(score_raw / 100, 4) if score_raw is not None else None
    nps_raw = r.get("nps", "")
    nps_v   = parse_nps(nps_raw)
    date_s  = parse_date_display(r.get("date", ""))
    trainer = r.get("pic", "")
    parts   = r.get("participants_num", 1) or 1
    trainee = r.get("trainee", "")

    # Exam Done?
    if method in ("Class", "Online"):
        exam_done = "Yes" if score_raw is not None else "No"
    else:
        exam_done = "NA"

    # Also Attended
    also_attended = ""
    if trainee == "FSM":
        also_attended = "Store Managers"
    elif trainee == "Promoter":
        also_attended = "Supervisor"

    row_vals = [
        idx,           # #
        date_s,        # Date
        r.get("city",""),   # City/Area
        method,        # Method
        trainee,       # Primary Audience
        also_attended, # Also Attended
        r.get("channel",""),  # Channel/Store
        topic,         # Training Topic (dropdown)
        model,         # Focus Model
        parts,         # # Participants
        1,             # Sessions Qty
        score_d,       # Exam Score
        nps_v,         # NPS /10
        trainer,       # Trainer PIC
        exam_done,     # Exam Done?
        "No",          # Repeated?
        None,          # Session Photos
        None,          # Exam Snapshots
    ]

    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_log.cell(row=row_num, column=col_idx)
        cell.value = val
        if col_idx == 1:
            # Row number — light gray
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        elif col_idx in (2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16):
            cell.fill = PatternFill("solid", fgColor="FFFDE7")  # light yellow

print(f"✓ Session Log: {len(apr)} rows written")

# ════════════════════════════════════════════════════════════════════════════
# 3.  EXAM RESULTS — write scored sessions
# ════════════════════════════════════════════════════════════════════════════
ws_exam = wb["📝 Exam Results"]
clear_rows(ws_exam, start_row=5, end_row=ws_exam.max_row)

# Use scored sessions; derive level from method/topic
exam_row = 5
exam_idx = 1
for r in apr:
    score_raw = r.get("score_num")
    if score_raw is None:
        continue
    score_d = round(score_raw / 100, 4)

    method  = map_method(r.get("method", "Class"))
    content = r.get("content", "")
    topic   = map_topic(content, method)
    date_s  = parse_date_display(r.get("date", ""))
    trainer = r.get("pic", "")
    city    = r.get("city", "")
    channel = r.get("channel", "")
    trainee = r.get("trainee", "")

    level_tested = topic

    # Pass/Fail: template threshold is 90%
    pass_fail = "Pass" if score_raw >= 90 else "Fail"

    row_vals = [
        exam_idx,      # #
        date_s,        # Date
        r.get("promo_name", trainee),  # Attendee Name
        "",            # Phone Number (not in source data)
        channel,       # Store/Channel
        city,          # City
        level_tested,  # Level Tested (dropdown)
        topic,         # Training Topic (dropdown)
        score_d,       # Score (decimal)
        None,          # Prev Score
        pass_fail,     # Pass/Fail
        "",            # Star Earned (CIPHER auto)
    ]

    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_exam.cell(row=exam_row, column=col_idx)
        cell.value = val
        if col_idx in (2, 3, 5, 6, 7, 8, 9, 10):
            cell.fill = PatternFill("solid", fgColor="FFFDE7")

    exam_row += 1
    exam_idx += 1

print(f"✓ Exam Results: {exam_idx - 1} rows written")

# ════════════════════════════════════════════════════════════════════════════
# 4.  NEXT MONTH PLAN — May 2026 sessions from Jun-dated rows
# ════════════════════════════════════════════════════════════════════════════
ws_plan = wb["📆 Next Month Plan"]
clear_rows(ws_plan, start_row=5, end_row=ws_plan.max_row)

plan_row = 5
plan_idx = 1

# Use up to 20 Jun rows as May plan entries
for raw in jun_rows[:20]:
    # raw row structure (17-23 cols):
    # [sr, cat, city, method, attend_type, date, trainee, channel, loc_or_promo,
    #  name_or_phone, content, model, score, nps, participants, pic, ...]
    if len(raw) < 16:
        continue
    city    = raw[2] if len(raw) > 2 else ""
    method  = map_method(raw[3] if len(raw) > 3 else "On-Site")
    trainee = raw[6] if len(raw) > 6 else ""
    channel = raw[7] if len(raw) > 7 else ""
    content = raw[10] if len(raw) > 10 else ""
    model   = raw[11] if len(raw) > 11 else ""
    parts   = int(raw[14]) if len(raw) > 14 and str(raw[14]).isdigit() else 1
    date_raw = raw[5] if len(raw) > 5 else ""
    date_may = date_raw.replace("Jun", "May") if date_raw else ""

    topic = map_topic(content, method)

    # Priority: Class = High, Online = Medium, Instore = Normal
    priority = {"Class": "High", "Online": "Medium", "Instore": "Normal"}.get(method, "Normal")

    row_vals = [
        plan_idx,      # #
        date_may,      # Planned Date
        city,          # City
        method,        # Method (dropdown)
        trainee,       # Primary Audience
        channel,       # Channel/Store
        topic,         # Training Topic (dropdown)
        model,         # Focus Model
        parts,         # Expected Participants
        priority,      # Priority (dropdown)
        "",            # Notes
    ]

    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_plan.cell(row=plan_row, column=col_idx)
        cell.value = val
        if col_idx in (2, 3, 4, 5, 6, 7, 8, 9, 10):
            cell.fill = PatternFill("solid", fgColor="FFFDE7")

    plan_row += 1
    plan_idx += 1

print(f"✓ Next Month Plan: {plan_idx - 1} rows written")

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"\n✅  Saved: {OUTPUT}")
print(f"   File size: {os.path.getsize(OUTPUT):,} bytes")
