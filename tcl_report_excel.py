#!/usr/bin/env python3
"""TCL Electronics KSA — Excel Report Generator
   Produces a fully formatted .xlsx report from tcl_training_data.json
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_FILE   = Path(__file__).parent / "tcl_training_data.json"
OUTPUT_FILE = Path(__file__).parent / f"TCL_Training_Report_{date.today().strftime('%Y_%m_%d')}.xlsx"

# ── Brand colours ─────────────────────────────────────────────────────────────
RED       = "C8102E"   # TCL red
DARK      = "1A1A2E"   # near-black
WHITE     = "FFFFFF"
LIGHT_RED = "FAD7DC"
LIGHT_GREY= "F5F5F5"
MID_GREY  = "D9D9D9"
GOLD_CLR  = "FFD700"
SILVER_CLR= "C0C0C0"
BRONZE_CLR= "CD7F32"

LEVEL_COLORS = {
    "Champion": "C8102E",
    "Gold":     "FFD700",
    "Silver":   "C0C0C0",
    "Bronze":   "CD7F32",
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK, size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border() -> Border:
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row: int, values: list, bg=RED, fg=WHITE) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=11)
        c.alignment = center()
        c.border = thin_border()

def apply_row(ws, row: int, values: list, bg=WHITE, bold=False) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=DARK)
        c.alignment = left()
        c.border = thin_border()

def set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def banner(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = title
    c.fill = fill(RED)
    c.font = font(bold=True, color=WHITE, size=16)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = subtitle
    c.fill = fill(DARK)
    c.font = font(bold=False, color=WHITE, size=10)
    c.alignment = center()
    ws.row_dimensions[2].height = 20


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    today = date.today().strftime("%d %B %Y")
    banner(ws, "TCL ELECTRONICS KSA — PROMOTER TRAINING REPORT",
           f"Generated: {today}  |  Period: {date.today().strftime('%B %Y')}")

    promoters = list(data["promoters"].values())
    comps     = list(data["competitions"].values())

    total   = len(promoters)
    avg_pts = round(sum(p["points"] for p in promoters) / total, 1) if total else 0
    levels  = {"Champion": 0, "Gold": 0, "Silver": 0, "Bronze": 0}
    for p in promoters:
        levels[p["level"]] = levels.get(p["level"], 0) + 1

    total_quizzes = sum(len(p.get("quiz_results", [])) for p in promoters)
    passed_quizzes = sum(
        sum(1 for r in p.get("quiz_results", []) if r["passed"])
        for p in promoters
    )

    # KPI cards (row 4+)
    ws.row_dimensions[3].height = 14
    kpis = [
        ("Total Promoters",    total,         LIGHT_RED),
        ("Avg Points",         avg_pts,        "DDEEFF"),
        ("Competitions Run",   len(comps),     "DDFFD6"),
        ("Quizzes Completed",  total_quizzes,  "FFF3CC"),
        ("Quizzes Passed",     passed_quizzes, "DDFFD6"),
    ]

    ws.merge_cells("A4:G4")
    ws["A4"].value = "KEY METRICS"
    ws["A4"].fill  = fill(DARK)
    ws["A4"].font  = font(bold=True, color=WHITE, size=12)
    ws["A4"].alignment = center()
    ws.row_dimensions[4].height = 24

    for col, (label, value, bg) in enumerate(kpis, 1):
        ws.merge_cells(
            start_row=5, start_column=col,
            end_row=5,   end_column=col
        )
        lc = ws.cell(row=5, column=col, value=label)
        lc.fill = fill(bg); lc.font = font(bold=True, size=9, color=DARK)
        lc.alignment = center(); lc.border = thin_border()

        vc = ws.cell(row=6, column=col, value=value)
        vc.fill = fill(bg); vc.font = font(bold=True, size=18, color=RED)
        vc.alignment = center(); vc.border = thin_border()
        ws.row_dimensions[6].height = 36

    # Level breakdown table
    ws.row_dimensions[7].height = 10
    ws.merge_cells("A8:G8")
    ws["A8"].value = "PROMOTER LEVELS"
    ws["A8"].fill  = fill(DARK)
    ws["A8"].font  = font(bold=True, color=WHITE, size=12)
    ws["A8"].alignment = center()
    ws.row_dimensions[8].height = 24

    apply_header(ws, 9, ["Level", "Count", "% of Team", "", "", "", ""])
    for i, (lvl, cnt) in enumerate(levels.items(), 10):
        pct = f"{cnt/total*100:.0f}%" if total else "0%"
        color = LEVEL_COLORS.get(lvl, WHITE)
        ws.cell(row=i, column=1, value=lvl).fill  = fill(color)
        ws.cell(row=i, column=1).font             = font(bold=True, color=DARK if lvl != "Champion" else WHITE)
        ws.cell(row=i, column=1).alignment        = center()
        ws.cell(row=i, column=1).border           = thin_border()
        ws.cell(row=i, column=2, value=cnt).fill  = fill(LIGHT_GREY)
        ws.cell(row=i, column=2).alignment        = center()
        ws.cell(row=i, column=2).border           = thin_border()
        ws.cell(row=i, column=3, value=pct).fill  = fill(LIGHT_GREY)
        ws.cell(row=i, column=3).alignment        = center()
        ws.cell(row=i, column=3).border           = thin_border()
        ws.row_dimensions[i].height = 22

    # Pie chart — level distribution
    pie = PieChart()
    pie.title = "Promoter Level Distribution"
    pie.style = 10
    labels = Reference(ws, min_col=1, min_row=10, max_row=13)
    data_ref = Reference(ws, min_col=2, min_row=9, max_row=13)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = None
    slice_colors = [BRONZE_CLR, SILVER_CLR, GOLD_CLR, RED]
    for i, color in enumerate(slice_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    pie.width  = 14
    pie.height = 10
    ws.add_chart(pie, "D9")

    set_col_widths(ws, [18, 10, 12, 14, 14, 14, 14])


def build_leaderboard(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Leaderboard")
    ws.sheet_view.showGridLines = False

    banner(ws, "PROMOTER LEADERBOARD",
           f"{date.today().strftime('%B %Y')} Rankings")

    headers = ["Rank", "Name", "Store", "Points", "Level", "Quizzes Passed", "Competitions"]
    apply_header(ws, 4, headers)
    ws.row_dimensions[4].height = 24

    ps = sorted(data["promoters"].values(), key=lambda x: x["points"], reverse=True)
    medals = {1: "🥇 1st", 2: "🥈 2nd", 3: "🥉 3rd"}

    for i, p in enumerate(ps, 1):
        row = i + 4
        bg = LIGHT_RED if i <= 3 else (LIGHT_GREY if i % 2 == 0 else WHITE)
        quizzes_passed = sum(1 for r in p.get("quiz_results", []) if r["passed"])
        comps_entered  = len(p.get("comp_entries", []))

        values = [
            medals.get(i, str(i)),
            p["name"],
            p["store"],
            p["points"],
            p["level"],
            quizzes_passed,
            comps_entered,
        ]
        apply_row(ws, row, values, bg=bg)

        # Colour the level cell
        lc = ws.cell(row=row, column=5)
        lc.fill = fill(LEVEL_COLORS.get(p["level"], WHITE))
        lc.font = font(bold=True, color=DARK)
        lc.alignment = center()

        # Bold top 3
        if i <= 3:
            for col in range(1, 8):
                ws.cell(row=row, column=col).font = font(bold=True, color=DARK)

        ws.row_dimensions[row].height = 22

    # Bar chart — points per promoter
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Points per Promoter"
    chart.style = 10
    chart.y_axis.title = "Points"
    chart.x_axis.title = "Promoter"
    chart.shape = 4

    last_row = len(ps) + 4
    data_ref  = Reference(ws, min_col=4, min_row=4, max_row=last_row)
    cats      = Reference(ws, min_col=2, min_row=5, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = RED
    chart.width  = 20
    chart.height = 12
    ws.add_chart(chart, f"A{last_row + 3}")

    set_col_widths(ws, [12, 22, 22, 10, 14, 16, 14])


def build_quiz_performance(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Quiz Performance")
    ws.sheet_view.showGridLines = False

    banner(ws, "QUIZ PERFORMANCE", "All promoter quiz results")

    headers = ["Promoter", "Store", "Quiz", "Score", "Total", "Pct %", "Status", "Date"]
    apply_header(ws, 4, headers)
    ws.row_dimensions[4].height = 24

    row = 5
    for p in data["promoters"].values():
        results = p.get("quiz_results", [])
        if not results:
            bg = LIGHT_GREY
            apply_row(ws, row, [p["name"], p["store"], "—", "—", "—", "—", "No attempts", "—"], bg=bg)
            ws.row_dimensions[row].height = 20
            row += 1
            continue
        for r in results:
            bg = "DDFFD6" if r["passed"] else LIGHT_RED
            values = [
                p["name"], p["store"],
                r["quiz"], r["score"], r["total"],
                f"{r['pct']}%",
                "PASS ✓" if r["passed"] else "FAIL ✗",
                r["date"],
            ]
            apply_row(ws, row, values, bg=bg)
            status_cell = ws.cell(row=row, column=7)
            status_cell.font = font(bold=True,
                                    color="1A7A3C" if r["passed"] else RED)
            status_cell.alignment = center()
            ws.row_dimensions[row].height = 20
            row += 1

    set_col_widths(ws, [22, 22, 16, 8, 8, 8, 12, 14])


def build_competition_results(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Competition Results")
    ws.sheet_view.showGridLines = False

    banner(ws, "COMPETITION RESULTS", "Rubric: Knowledge 35% · Creativity 25% · Selling 40%")

    row = 4
    for comp in data["competitions"].values():
        if not comp["entries"]:
            continue

        # Competition title row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"{comp['name']}  |  {comp['date']}")
        c.fill = fill(DARK); c.font = font(bold=True, color=WHITE, size=12)
        c.alignment = center()
        ws.row_dimensions[row].height = 26
        row += 1

        headers = ["Rank", "Promoter", "Store", "Knowledge /4",
                   "Creativity /4", "Selling /4", "Score /4", "Score %"]
        apply_header(ws, row, headers)
        ws.row_dimensions[row].height = 24
        row += 1

        ranked = sorted(comp["entries"], key=lambda e: e["weighted_score"], reverse=True)
        medals = {1: "🥇 1st", 2: "🥈 2nd", 3: "🥉 3rd"}

        for i, e in enumerate(ranked, 1):
            bg = LIGHT_RED if i <= 3 else (LIGHT_GREY if i % 2 == 0 else WHITE)
            s = e["scores"]
            values = [
                medals.get(i, str(i)),
                e["promoter_name"], e["store"],
                s["knowledge"], s["creativity"], s["selling"],
                round(e["weighted_score"], 3),
                f"{e['pct']}%",
            ]
            apply_row(ws, row, values, bg=bg)
            if i <= 3:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).font = font(bold=True)
            ws.row_dimensions[row].height = 22
            row += 1

        row += 2  # gap between competitions

    set_col_widths(ws, [12, 22, 22, 14, 14, 12, 12, 10])


def build_promoter_profiles(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Promoter Profiles")
    ws.sheet_view.showGridLines = False

    banner(ws, "PROMOTER PROFILES", "Individual history and points log")

    row = 4
    for p in data["promoters"].values():
        # Profile header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        lvl_color = LEVEL_COLORS.get(p["level"], WHITE)
        c = ws.cell(row=row, column=1,
                    value=f"{p['id']}  |  {p['name']}  |  {p['store']}  |  {p['level']} — {p['points']} pts")
        c.fill = fill(lvl_color)
        c.font = font(bold=True, color=DARK, size=12)
        c.alignment = left()
        ws.row_dimensions[row].height = 26
        row += 1

        # Points history
        history = p.get("history", [])
        if history:
            apply_header(ws, row, ["Date", "Points", "Reason", "", "", ""], bg=DARK)
            ws.row_dimensions[row].height = 20
            row += 1
            for h in history:
                sign = f"+{h['pts']}" if h["pts"] >= 0 else str(h["pts"])
                bg = "DDFFD6" if h["pts"] > 0 else LIGHT_RED
                apply_row(ws, row, [h["date"], sign, h["reason"], "", "", ""], bg=bg)
                ws.row_dimensions[row].height = 18
                row += 1
        else:
            ws.cell(row=row, column=1, value="No activity yet").font = font(italic=True, color="888888")
            row += 1

        row += 2  # gap

    set_col_widths(ws, [14, 10, 40, 14, 14, 14])


def build_training_plan(wb: Workbook) -> None:
    plan_file = Path(__file__).parent / "training_data.json"
    if not plan_file.exists():
        return

    plan = json.loads(plan_file.read_text())
    ws = wb.create_sheet("June Training Plan")
    ws.sheet_view.showGridLines = False

    banner(ws, "JUNE 2026 TRAINING PLAN", "Personal monthly training tracker")

    tasks = [t for w in plan["weeks"] for t in w["tasks"]]
    done  = sum(1 for t in tasks if t["status"] == "done")
    total = len(tasks)
    pct   = int(done / total * 100) if total else 0

    ws.merge_cells("A4:G4")
    c = ws["A4"]
    c.value = f"Progress: {done} / {total} tasks complete  ({pct}%)"
    c.fill  = fill(DARK)
    c.font  = font(bold=True, color=WHITE, size=12)
    c.alignment = center()
    ws.row_dimensions[4].height = 24

    row = 6
    for week in plan["weeks"]:
        # Week header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"Week {week['week']}: {week['label']}")
        c.fill = fill(RED); c.font = font(bold=True, color=WHITE)
        c.alignment = left()
        ws.row_dimensions[row].height = 22
        row += 1

        apply_header(ws, row, ["ID", "Day", "Task", "Status", ""], bg=DARK)
        ws.row_dimensions[row].height = 20
        row += 1

        for t in week["tasks"]:
            status = t["status"]
            bg = "DDFFD6" if status == "done" else (LIGHT_GREY if status == "skip" else WHITE)
            icon = {"done": "✓ Done", "todo": "○ To Do", "skip": "— Skip"}.get(status, status)
            apply_row(ws, row, [t["id"], t["day"], t["name"], icon, ""], bg=bg)
            sc = ws.cell(row=row, column=4)
            sc.font = font(bold=True,
                           color="1A7A3C" if status == "done" else
                                 ("888888" if status == "skip" else DARK))
            sc.alignment = center()
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1

    set_col_widths(ws, [8, 14, 45, 12, 10])


# ── Main ──────────────────────────────────────────────────────────────────────

def generate() -> None:
    if not DATA_FILE.exists():
        print("No training data found. Run tcl_training.py first to register promoters.")
        return

    data = json.loads(DATA_FILE.read_text())

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_summary(wb, data)
    build_leaderboard(wb, data)
    build_quiz_performance(wb, data)
    build_competition_results(wb, data)
    build_promoter_profiles(wb, data)
    build_training_plan(wb)

    wb.save(OUTPUT_FILE)
    print(f"Report saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    generate()
