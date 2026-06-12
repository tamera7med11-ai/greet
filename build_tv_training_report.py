#!/usr/bin/env python3
"""
TCL TV Training Report — Jan to May 2026
Reads parsed training data and generates a fully formatted Excel workbook.
"""

import json, re, pickle
from collections import defaultdict
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule, CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / f"TCL_TV_Training_Jan_May_2026.xlsx"
DATA   = Path(__file__).parent / "tv_training_parsed.pkl"

# ── Colours ───────────────────────────────────────────────────────────────────
RED      = "C8102E"; DARK     = "1A1A2E"; WHITE    = "FFFFFF"
LGREY    = "F5F5F5"; MGREY    = "D9D9D9"; YELLOW   = "FFF3CC"
LGREEN   = "DDFFD6"; GREEN    = "1A7A3C"; LRED     = "FAD7DC"
BLUE     = "1F4E79"; LBLUE    = "D6E4F0"; ORANGE   = "E8630A"

def fill(c):  return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=DARK, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def ctr():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft():  return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def bdr(c=MGREY):
    s = Side(style="thin", color=c); return Border(left=s,right=s,top=s,bottom=s)

def hdr(ws, row, cols, bg=RED, fg=WHITE, h=26):
    for i, v in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill=fill(bg); c.font=fnt(True,fg); c.alignment=ctr(); c.border=bdr()
    ws.row_dimensions[row].height = h

def banner(ws, title, sub, end_col="H"):
    ws.merge_cells(f"A1:{end_col}1")
    c=ws["A1"]; c.value=title; c.fill=fill(RED)
    c.font=fnt(True,WHITE,16); c.alignment=ctr(); ws.row_dimensions[1].height=40
    ws.merge_cells(f"A2:{end_col}2")
    c=ws["A2"]; c.value=sub; c.fill=fill(DARK)
    c.font=fnt(False,WHITE,10); c.alignment=ctr(); ws.row_dimensions[2].height=20

def col_w(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def kpi_card(ws, row, col, label, value, bg=LBLUE):
    lc = ws.cell(row=row,   column=col, value=label)
    lc.fill=fill(bg); lc.font=fnt(True,DARK,9); lc.alignment=ctr(); lc.border=bdr()
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.fill=fill(bg); vc.font=fnt(True,RED,20); vc.alignment=ctr(); vc.border=bdr()
    ws.row_dimensions[row+1].height = 40

# ── Parse data ────────────────────────────────────────────────────────────────

def parse_data():
    with open('/root/.claude/projects/-home-user-greet/82218592-4f36-5112-b730-0955daaf23f9/tool-results/mcp-ca1b9ec8-582c-41dd-b292-9e6b4c41be6a-read_file_content-1781282115949.txt') as f:
        content = json.load(f)['fileContent']

    records = []
    for line in content.split('\n'):
        line = line.strip()
        if not line.startswith('|') or ':-:' in line:
            continue
        cells = [c.strip().replace('&#10;',' ').replace('\\#','#') for c in line.split('|')[1:-1]]

        # Handle leading empty cell (May sheet has offset)
        if cells and cells[0] == '':
            cells = cells[1:]

        if not cells or not re.match(r'^\d+$', cells[0].strip()):
            continue

        # Normalise to standard columns based on length
        # Standard 16-col: Sr|Cat|City|Method|Attend|Date|Type|Channel|Map|PromoName|Content|Model|Score|NPS|Participants|PIC
        # 17-col (Apr/May): Sr|Cat|City|Method|Attend|Date|Type|Channel|Map|SubChannel|PromoName|Content|Model|Score|NPS|Participants|PIC
        n = len(cells)
        if n >= 17:
            rec = dict(
                sr=cells[0], city=cells[2], method=cells[3], attend_type=cells[4],
                date=cells[5], trainee=cells[6], channel=cells[7],
                promo_name=cells[10], content=cells[11], model=cells[12],
                score=cells[13], nps=cells[14],
                participants=cells[15] if n>15 else '1', pic=cells[16] if n>16 else ''
            )
        elif n >= 16:
            rec = dict(
                sr=cells[0], city=cells[2], method=cells[3], attend_type=cells[4],
                date=cells[5], trainee=cells[6], channel=cells[7],
                promo_name=cells[9], content=cells[10], model=cells[11],
                score=cells[12], nps=cells[13],
                participants=cells[14] if n>14 else '1', pic=cells[15] if n>15 else ''
            )
        else:
            continue

        # Detect month
        d = rec['date']
        rec['month'] = next((m for m in ['Jan','Feb','Mar','Apr','May'] if m in d), 'Other')

        # Parse score
        s = rec['score'].replace('%','').strip()
        try:    rec['score_num'] = float(s)
        except: rec['score_num'] = None

        # Parse participants
        try:    rec['participants_num'] = int(rec['participants'])
        except: rec['participants_num'] = 1

        records.append(rec)

    return records

# ── Sheet 1: Executive Summary ─────────────────────────────────────────────────

def sheet_summary(wb, records):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    banner(ws, "TCL TV TRAINING REPORT — KSA | JAN–MAY 2026",
           f"Trainer: Tamer Ahmed  |  Generated: {date.today().strftime('%d %B %Y')}", "I")

    months = ['Jan','Feb','Mar','Apr','May']

    # KPI cards row 4-5
    ws.merge_cells("A3:I3")
    ws["A3"].value = "KEY PERFORMANCE INDICATORS"
    ws["A3"].fill = fill(DARK); ws["A3"].font = fnt(True,WHITE,12); ws["A3"].alignment=ctr()
    ws.row_dimensions[3].height = 26

    total_sessions     = len(records)
    total_participants = sum(r['participants_num'] for r in records)
    scored             = [r['score_num'] for r in records if r['score_num'] is not None]
    avg_score          = round(sum(scored)/len(scored),1) if scored else 0
    cities             = len(set(r['city'] for r in records if r['city']))
    promoters          = sum(1 for r in records if 'Promoter' in r['trainee'])
    fsm                = sum(1 for r in records if 'FSM' in r['trainee'])

    kpis = [
        ("Total Sessions",    total_sessions,     LBLUE),
        ("Total Participants",total_participants,  LGREEN),
        ("Avg Exam Score",    f"{avg_score}%",    YELLOW),
        ("Cities Covered",    cities,             LBLUE),
        ("Promoter Sessions", promoters,          LGREEN),
        ("FSM Sessions",      fsm,                LBLUE),
    ]
    for i, (label, val, bg) in enumerate(kpis, 1):
        kpi_card(ws, 4, i, label, val, bg)

    ws.row_dimensions[4].height = 22

    # Monthly breakdown table (row 7+)
    ws.merge_cells("A7:I7")
    ws["A7"].value = "MONTHLY BREAKDOWN"
    ws["A7"].fill=fill(DARK); ws["A7"].font=fnt(True,WHITE,12); ws["A7"].alignment=ctr()
    ws.row_dimensions[7].height = 26

    hdr(ws, 8, ["Month","Sessions","Participants","Scored Sessions","Avg Score","Best Score",
                "On-Site","Online / Class","Cities"], bg=RED)

    monthly_stats = []
    for m in months:
        rs = [r for r in records if r['month'] == m]
        if not rs: continue
        parts    = sum(r['participants_num'] for r in rs)
        s_nums   = [r['score_num'] for r in rs if r['score_num'] is not None]
        avg_s    = round(sum(s_nums)/len(s_nums),1) if s_nums else 0
        best_s   = max(s_nums) if s_nums else 0
        onsite   = sum(1 for r in rs if 'site' in r['method'].lower() or 'Site' in r['method'])
        other    = len(rs) - onsite
        cities_m = len(set(r['city'] for r in rs if r['city']))
        monthly_stats.append((m, len(rs), parts, len(s_nums), avg_s, best_s, onsite, other, cities_m))

    for i, row_data in enumerate(monthly_stats):
        row = i + 9
        bg  = LGREY if i % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill=fill(bg); c.alignment=ctr(); c.border=bdr()
            if col == 1: c.font=fnt(True,RED)
            elif col == 5:
                c.number_format = '0.0"%"'
                c.font=fnt(True,GREEN if val>=85 else ORANGE if val>=70 else RED)
        ws.row_dimensions[row].height = 22

    # Totals row
    row = len(monthly_stats) + 9
    totals = ["TOTAL",
              sum(x[1] for x in monthly_stats),
              sum(x[2] for x in monthly_stats),
              sum(x[3] for x in monthly_stats),
              round(sum(x[4] for x in monthly_stats if x[4]>0)/max(1,sum(1 for x in monthly_stats if x[4]>0)),1),
              max((x[5] for x in monthly_stats),default=0),
              sum(x[6] for x in monthly_stats),
              sum(x[7] for x in monthly_stats),
              ""]
    for col, val in enumerate(totals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill=fill(RED); c.font=fnt(True,WHITE); c.alignment=ctr(); c.border=bdr()
    ws.row_dimensions[row].height = 24

    # Bar chart — sessions per month
    chart = BarChart()
    chart.type="col"; chart.title="Sessions per Month"; chart.style=10
    chart.y_axis.title="Sessions"; chart.x_axis.title="Month"
    data_ref = Reference(ws, min_col=2, min_row=8, max_row=row-1)
    cats     = Reference(ws, min_col=1, min_row=9, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = RED
    chart.width=14; chart.height=10
    ws.add_chart(chart, "A14")

    # Line chart — avg score trend
    line = LineChart()
    line.title="Avg Score Trend (%)"; line.style=10
    line.y_axis.title="Score %"; line.y_axis.numFmt='0'
    s_ref = Reference(ws, min_col=5, min_row=8, max_row=row-1)
    line.add_data(s_ref, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties.line.solidFill = RED
    line.width=14; line.height=10
    ws.add_chart(line, "F14")

    col_w(ws, [12,12,14,16,12,12,12,14,12])


# ── Sheet 2: All Sessions ──────────────────────────────────────────────────────

def sheet_all_sessions(wb, records):
    ws = wb.create_sheet("All Sessions")
    ws.sheet_view.showGridLines = False
    banner(ws, "ALL TRAINING SESSIONS — JAN TO MAY 2026", "Complete session log", "N")

    cols = ["#","Month","Date","City","Method","Trainee Type","Channel",
            "Promoter / Trainee","Content","Model","Score","NPS","Participants","PIC"]
    hdr(ws, 4, cols)

    month_colors = {"Jan":LBLUE,"Feb":LGREEN,"Mar":YELLOW,"Apr":"FFE4CC","May":LGREY}

    for i, r in enumerate(records, 1):
        row = i + 4
        bg  = month_colors.get(r['month'], WHITE)
        vals = [i, r['month'], r['date'], r['city'], r['method'], r['trainee'],
                r['channel'], r['promo_name'], r['content'], r['model'],
                r['score'], r['nps'], r['participants_num'], r['pic']]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill=fill(bg); c.alignment=lft() if col>3 else ctr()
            c.border=bdr(); c.font=fnt(size=9)
        # Colour score cell
        sc = ws.cell(row=row, column=11)
        if r['score_num'] is not None:
            if   r['score_num'] >= 90: sc.fill=fill(LGREEN); sc.font=fnt(True,GREEN,9)
            elif r['score_num'] >= 75: sc.fill=fill(YELLOW);  sc.font=fnt(True,ORANGE,9)
            else:                      sc.fill=fill(LRED);    sc.font=fnt(True,RED,9)
        ws.row_dimensions[row].height = 18

    col_w(ws, [5,8,10,12,10,14,22,22,22,16,10,8,12,10])


# ── Sheet 3: Monthly Detail (one per month) ───────────────────────────────────

def sheet_monthly(wb, records, month, color):
    ws = wb.create_sheet(f"{month} Detail")
    ws.sheet_view.showGridLines = False
    banner(ws, f"TCL TV TRAINING — {month.upper()} 2026",
           f"Detailed session log for {month} 2026", "L")

    rs = [r for r in records if r['month'] == month]
    if not rs:
        ws["A4"].value = "No data for this month."
        return

    # Month KPIs
    parts  = sum(r['participants_num'] for r in rs)
    scored = [r['score_num'] for r in rs if r['score_num'] is not None]
    avg_s  = round(sum(scored)/len(scored),1) if scored else 0

    for col, (label, val) in enumerate([
        ("Sessions", len(rs)), ("Participants", parts),
        ("Scored", len(scored)), (f"Avg Score", f"{avg_s}%")
    ], 1):
        kpi_card(ws, 4, col, label, val, color)
    ws.row_dimensions[4].height = 22

    hdr(ws, 6, ["#","Date","City","Method","Type","Channel","Trainee/Promoter",
                "Content","Model","Score","Participants","PIC"], bg=RED)

    for i, r in enumerate(rs, 1):
        row = i + 6
        bg  = LGREY if i%2==0 else WHITE
        vals = [i, r['date'], r['city'], r['method'], r['trainee'], r['channel'],
                r['promo_name'], r['content'], r['model'], r['score'],
                r['participants_num'], r['pic']]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill=fill(bg); c.border=bdr()
            c.alignment=lft() if col>2 else ctr(); c.font=fnt(size=10)
        # Score colour
        sc = ws.cell(row=row, column=10)
        if r['score_num'] is not None:
            if   r['score_num']>=90: sc.fill=fill(LGREEN); sc.font=fnt(True,GREEN,10)
            elif r['score_num']>=75: sc.fill=fill(YELLOW);  sc.font=fnt(True,ORANGE,10)
            else:                    sc.fill=fill(LRED);    sc.font=fnt(True,RED,10)
        ws.row_dimensions[row].height = 20

    col_w(ws, [5,10,12,12,14,24,24,24,16,10,12,10])


# ── Sheet 4: Trainer Performance ──────────────────────────────────────────────

def sheet_trainers(wb, records):
    ws = wb.create_sheet("Trainer Performance")
    ws.sheet_view.showGridLines = False
    banner(ws, "TRAINER PERFORMANCE OVERVIEW", "Sessions, reach, and scores per PIC", "H")

    # Gather per-trainer stats
    trainer_data = defaultdict(lambda: defaultdict(list))
    for r in records:
        pic = r['pic'].strip()
        if not pic or pic in ('','Unknown') or pic.isdigit():
            pic = 'Unknown'
        trainer_data[pic][r['month']].append(r)

    hdr(ws, 4, ["Trainer","Jan","Feb","Mar","Apr","May","Total Sessions","Total Participants","Avg Score"])

    months = ['Jan','Feb','Mar','Apr','May']
    row = 5
    totals = defaultdict(int)
    for trainer, mdata in sorted(trainer_data.items(), key=lambda x: -sum(len(v) for v in x[1].values())):
        bg = LGREY if row%2==0 else WHITE
        ws.cell(row=row, column=1, value=trainer).fill=fill(bg)
        ws.cell(row=row, column=1).font=fnt(True,RED)
        ws.cell(row=row, column=1).border=bdr(); ws.cell(row=row, column=1).alignment=lft()

        total_sess = 0; total_parts = 0; all_scores = []
        for mi, m in enumerate(months, 2):
            rs_m = mdata.get(m, [])
            n = len(rs_m)
            total_sess  += n
            total_parts += sum(r['participants_num'] for r in rs_m)
            all_scores  += [r['score_num'] for r in rs_m if r['score_num'] is not None]
            c = ws.cell(row=row, column=mi, value=n if n>0 else "—")
            c.fill=fill(LBLUE if n>0 else bg); c.alignment=ctr(); c.border=bdr()
            c.font=fnt(True if n>0 else False, DARK)

        avg_s = round(sum(all_scores)/len(all_scores),1) if all_scores else "N/A"
        for col, val in [(7, total_sess),(8, total_parts),(9, f"{avg_s}%" if isinstance(avg_s,float) else avg_s)]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill=fill(bg); c.alignment=ctr(); c.border=bdr(); c.font=fnt(True)

        ws.row_dimensions[row].height = 22
        row += 1

    col_w(ws, [16,10,10,10,10,10,14,16,12])


# ── Sheet 5: City & Channel Analysis ──────────────────────────────────────────

def sheet_cities(wb, records):
    ws = wb.create_sheet("City & Channel")
    ws.sheet_view.showGridLines = False
    banner(ws, "CITY & CHANNEL ANALYSIS", "Training reach by location and channel", "H")

    # City stats
    ws.merge_cells("A4:H4")
    ws["A4"].value="BY CITY"; ws["A4"].fill=fill(DARK)
    ws["A4"].font=fnt(True,WHITE,12); ws["A4"].alignment=ctr()
    ws.row_dimensions[4].height = 24

    hdr(ws, 5, ["City","Jan","Feb","Mar","Apr","May","Total","Participants"])
    months = ['Jan','Feb','Mar','Apr','May']
    city_data = defaultdict(lambda: defaultdict(list))
    for r in records:
        if r['city']: city_data[r['city']][r['month']].append(r)

    row = 6
    for city, mdata in sorted(city_data.items(), key=lambda x:-sum(len(v) for v in x[1].values())):
        bg = LGREY if row%2==0 else WHITE
        ws.cell(row=row,column=1,value=city).fill=fill(bg)
        ws.cell(row=row,column=1).font=fnt(True,DARK); ws.cell(row=row,column=1).border=bdr(); ws.cell(row=row,column=1).alignment=lft()
        total=0; total_p=0
        for mi,m in enumerate(months,2):
            rs=mdata.get(m,[]); n=len(rs)
            total+=n; total_p+=sum(r['participants_num'] for r in rs)
            c=ws.cell(row=row,column=mi,value=n if n else "—")
            c.fill=fill(LBLUE if n>0 else bg); c.alignment=ctr(); c.border=bdr()
        ws.cell(row=row,column=7,value=total).fill=fill(bg); ws.cell(row=row,column=7).font=fnt(True,RED); ws.cell(row=row,column=7).alignment=ctr(); ws.cell(row=row,column=7).border=bdr()
        ws.cell(row=row,column=8,value=total_p).fill=fill(bg); ws.cell(row=row,column=8).alignment=ctr(); ws.cell(row=row,column=8).border=bdr()
        ws.row_dimensions[row].height=20; row+=1

    # Top channels table
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].value="TOP 15 CHANNELS BY SESSIONS"; ws[f"A{row}"].fill=fill(DARK)
    ws[f"A{row}"].font=fnt(True,WHITE,12); ws[f"A{row}"].alignment=ctr(); ws.row_dimensions[row].height=24; row+=1

    channel_data = defaultdict(list)
    for r in records:
        if r['channel']: channel_data[r['channel']].append(r)

    hdr(ws, row, ["Channel","City","Sessions","Participants","Avg Score","Best Score","Trainee Types","Trainer"]); row+=1
    for ch, rs in sorted(channel_data.items(), key=lambda x:-len(x[1]))[:15]:
        bg = LGREY if row%2==0 else WHITE
        parts = sum(r['participants_num'] for r in rs)
        scores = [r['score_num'] for r in rs if r['score_num'] is not None]
        avg_s = f"{round(sum(scores)/len(scores),1)}%" if scores else "N/A"
        best_s = f"{max(scores)}%" if scores else "N/A"
        types = ", ".join(set(r['trainee'] for r in rs if r['trainee']))
        pics  = ", ".join(set(r['pic'] for r in rs if r['pic'] and not r['pic'].isdigit()))
        city  = rs[0]['city'] if rs else ''
        vals  = [ch, city, len(rs), parts, avg_s, best_s, types, pics]
        for col, val in enumerate(vals,1):
            c=ws.cell(row=row,column=col,value=val)
            c.fill=fill(bg); c.alignment=lft() if col in(1,7,8) else ctr()
            c.border=bdr(); c.font=fnt(size=10)
        ws.row_dimensions[row].height=20; row+=1

    col_w(ws, [20,12,10,12,12,12,20,14])


# ── Sheet 6: Score Analysis ────────────────────────────────────────────────────

def sheet_scores(wb, records):
    ws = wb.create_sheet("Score Analysis")
    ws.sheet_view.showGridLines = False
    banner(ws, "EXAM SCORE ANALYSIS", "Performance distribution and trends Jan–May 2026", "H")

    scored = [r for r in records if r['score_num'] is not None]
    months = ['Jan','Feb','Mar','Apr','May']

    # Score distribution buckets
    def bucket(s):
        if s>=90: return "Excellent (90–100%)"
        if s>=80: return "Good (80–89%)"
        if s>=70: return "Pass (70–79%)"
        return "Below Pass (<70%)"

    # Table — score distribution per month
    ws.merge_cells("A4:H4")
    ws["A4"].value="SCORE DISTRIBUTION BY MONTH"; ws["A4"].fill=fill(DARK)
    ws["A4"].font=fnt(True,WHITE,12); ws["A4"].alignment=ctr(); ws.row_dimensions[4].height=24

    buckets_order = ["Excellent (90–100%)","Good (80–89%)","Pass (70–79%)","Below Pass (<70%)"]
    bkt_colors    = {buckets_order[0]:LGREEN, buckets_order[1]:YELLOW, buckets_order[2]:"FFE4CC", buckets_order[3]:LRED}

    hdr(ws, 5, ["Score Band","Jan","Feb","Mar","Apr","May","Total","% of Scored"])
    row=6
    bkt_data = defaultdict(lambda: defaultdict(int))
    for r in scored:
        bkt_data[bucket(r['score_num'])][r['month']] += 1

    for bkt in buckets_order:
        bg = bkt_colors[bkt]
        ws.cell(row=row,column=1,value=bkt).fill=fill(bg); ws.cell(row=row,column=1).font=fnt(True); ws.cell(row=row,column=1).border=bdr(); ws.cell(row=row,column=1).alignment=lft()
        total=0
        for mi,m in enumerate(months,2):
            n=bkt_data[bkt].get(m,0); total+=n
            c=ws.cell(row=row,column=mi,value=n if n else "—")
            c.fill=fill(bg); c.alignment=ctr(); c.border=bdr()
        pct=f"{round(total/len(scored)*100,1)}%" if scored else "0%"
        ws.cell(row=row,column=7,value=total).fill=fill(bg); ws.cell(row=row,column=7).font=fnt(True); ws.cell(row=row,column=7).alignment=ctr(); ws.cell(row=row,column=7).border=bdr()
        ws.cell(row=row,column=8,value=pct).fill=fill(bg); ws.cell(row=row,column=8).alignment=ctr(); ws.cell(row=row,column=8).border=bdr()
        ws.row_dimensions[row].height=22; row+=1

    # Monthly avg score table (for chart)
    row+=2
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].value="MONTHLY AVG SCORE (FOR CHART)"; ws[f"A{row}"].fill=fill(DARK)
    ws[f"A{row}"].font=fnt(True,WHITE,12); ws[f"A{row}"].alignment=ctr(); ws.row_dimensions[row].height=24
    chart_start=row; row+=1
    hdr(ws,row,["Month","Avg Score","Sessions Scored","Min","Max"]); row+=1
    for m in months:
        rs_m=[r for r in scored if r['month']==m]
        if not rs_m:
            for col in range(1,6): ws.cell(row=row,column=col,value=m if col==1 else "N/A").border=bdr()
        else:
            vals=[m, round(sum(r['score_num'] for r in rs_m)/len(rs_m),1), len(rs_m), min(r['score_num'] for r in rs_m), max(r['score_num'] for r in rs_m)]
            for col,val in enumerate(vals,1):
                c=ws.cell(row=row,column=col,value=val); c.border=bdr()
                c.fill=fill(LGREEN if isinstance(val,float) and val>=85 else YELLOW if isinstance(val,float) and val>=70 else LGREY)
                c.alignment=ctr()
        ws.row_dimensions[row].height=22; row+=1

    # Bar chart
    chart=BarChart(); chart.type="col"; chart.title="Avg Score by Month"; chart.style=10
    chart.y_axis.title="Score %"; chart.y_axis.numFmt='0'
    d=Reference(ws,min_col=2,min_row=chart_start+1,max_row=chart_start+1+len(months))
    cats=Reference(ws,min_col=1,min_row=chart_start+2,max_row=chart_start+1+len(months))
    chart.add_data(d,titles_from_data=True); chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill=RED
    chart.width=16; chart.height=12
    ws.add_chart(chart,f"F{chart_start}")

    col_w(ws,[22,10,10,10,10,10,10,10])


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("Parsing data...")
    records = parse_data()
    print(f"  {len(records)} sessions loaded")

    wb = Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    sheet_summary(wb, records)
    sheet_all_sessions(wb, records)

    month_colors = {"Jan":LBLUE,"Feb":LGREEN,"Mar":YELLOW,"Apr":"FFE4CC","May":LGREY}
    for m, c in month_colors.items():
        sheet_monthly(wb, records, m, c)

    sheet_trainers(wb, records)
    sheet_cities(wb, records)
    sheet_scores(wb, records)

    wb.active = wb["Executive Summary"]
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

if __name__ == "__main__":
    main()
